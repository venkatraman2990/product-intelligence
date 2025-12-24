"""Excel exporter for contract extraction data.

Exports ContractData to a single flat Excel sheet with one row per contract.
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.schema import ContractData

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export extracted contract data to Excel format.

    Creates a single flat sheet with all fields as columns and
    one row per contract for easy analysis.
    """

    # Accelerant blue color (from design system)
    HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def __init__(self):
        """Initialize the Excel exporter."""
        pass

    def export(
        self,
        contracts: list[ContractData],
        output_path: str | Path,
        sheet_name: str = "Contract Extractions",
    ) -> Path:
        """Export contract data to an Excel file.

        Args:
            contracts: List of ContractData objects to export
            output_path: Path for the output Excel file
            sheet_name: Name for the worksheet

        Returns:
            Path to the created Excel file
        """
        output_path = Path(output_path)

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Exporting {len(contracts)} contracts to {output_path}")

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Get column headers from schema
        columns = ContractData.get_flat_columns()

        # Write header row
        self._write_header(ws, columns)

        # Write data rows
        for row_idx, contract in enumerate(contracts, start=2):
            flat_data = contract.to_flat_dict()
            self._write_row(ws, row_idx, columns, flat_data)

        # Auto-fit column widths
        self._auto_fit_columns(ws, columns)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add metadata sheet with export info
        self._add_metadata_sheet(wb, len(contracts))

        # Save workbook
        wb.save(output_path)
        logger.info(f"Successfully exported to {output_path}")

        return output_path

    def _write_header(self, ws, columns: list[str]):
        """Write styled header row.

        Args:
            ws: Worksheet object
            columns: List of column names
        """
        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=self._format_header(column_name))
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _write_row(self, ws, row_idx: int, columns: list[str], data: dict):
        """Write a data row.

        Args:
            ws: Worksheet object
            row_idx: Row number (1-indexed)
            columns: List of column names
            data: Flattened contract data dictionary
        """
        for col_idx, column_name in enumerate(columns, start=1):
            value = data.get(column_name)

            # Handle None values
            if value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self.BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _auto_fit_columns(self, ws, columns: list[str]):
        """Auto-fit column widths based on content.

        Args:
            ws: Worksheet object
            columns: List of column names
        """
        for col_idx, column_name in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Calculate max width from header and first 50 rows
            max_length = len(self._format_header(column_name))

            for row in range(2, min(52, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            # Cap column width and add padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_header(self, column_name: str) -> str:
        """Format column name for display.

        Args:
            column_name: Raw column name (e.g., "limits_per_risk")

        Returns:
            Formatted header (e.g., "Limits Per Risk")
        """
        return column_name.replace("_", " ").title()

    def _add_metadata_sheet(self, wb: Workbook, record_count: int):
        """Add a metadata sheet with export information.

        Args:
            wb: Workbook object
            record_count: Number of contracts exported
        """
        ws = wb.create_sheet("Export Info")

        metadata = [
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Records", record_count),
            ("Generated By", "Product Intelligence - Contract Extractor"),
            ("Version", "0.1.0"),
        ]

        for row_idx, (label, value) in enumerate(metadata, start=1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40


def export_to_excel(
    contracts: list[ContractData],
    output_path: str | Path,
) -> Path:
    """Convenience function to export contracts to Excel.

    Args:
        contracts: List of ContractData objects
        output_path: Output file path

    Returns:
        Path to created file
    """
    exporter = ExcelExporter()
    return exporter.export(contracts, output_path)
